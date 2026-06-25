# excel_logger.py
# Updated: 'Live Pickups' renamed to 'Item Events'
#          'Session Summaries' renamed to 'Purchase History'
# Also logs 'returned' events so the vendor can see corrections.

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

EXCEL_FILE = "smart_inventory_log.xlsx"

EVENT_HEADERS = [
    "Event ID", "Timestamp", "Customer Name",
    "Staff ID", "Item", "Action",
    "Unit Price (N)", "Signed Amount (N)", "Running Total (N)"
]

SUMMARY_HEADERS = [
    "Session ID", "Start Time", "End Time",
    "Customer Name", "Staff ID", "Items Purchased", "Session Total (N)"
]

def create_excel_file():
    wb = openpyxl.Workbook()

    # Sheet 1 — Item Events (was Live Pickups)
    ws1 = wb.active
    ws1.title = "Item Events"
    for col, header in enumerate(EVENT_HEADERS, start=1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
        cell.alignment = Alignment(horizontal="center")
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 16
    ws1.column_dimensions['H'].width = 16
    ws1.column_dimensions['I'].width = 18
    ws1.freeze_panes = "A2"

    # Sheet 2 — Purchase History (was Session Summaries)
    ws2 = wb.create_sheet(title="Purchase History")
    for col, header in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor="e94560")
        cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 48
    ws2.column_dimensions['G'].width = 18
    ws2.freeze_panes = "A2"

    wb.save(EXCEL_FILE)
    print(f"Excel file created: {EXCEL_FILE}")


def log_item_event(record: dict):
    """
    Logs a single item event (taken or returned) to the Item Events sheet.
    record: dict returned by LiveSession.process_event()
    """
    if not os.path.exists(EXCEL_FILE):
        create_excel_file()

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Item Events"]
    next_row = ws.max_row + 1
    event_id = f"EVT{next_row - 1:04d}"

    # Green row for taken, amber row for returned
    direction = record.get("direction", "taken")
    if direction == "returned":
        row_colour = "FFF3CD"   # amber — returned item
    else:
        row_colour = "e8e8f5" if next_row % 2 == 0 else "ffffff"

    row_fill = PatternFill("solid", fgColor=row_colour)

    row_data = [
        event_id,
        record["timestamp"],
        record["customer_name"],
        record["staff_id"],
        record["item"],
        record["direction"].upper(),
        record["unit_price"],
        record["line_total"],
        record["running_total"]
    ]

    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left")
        cell.font = Font(size=10)

    ws.cell(row=next_row, column=9).font = Font(bold=True, size=10)
    wb.save(EXCEL_FILE)
    print(f"Event logged: {record['item']} {record['direction']} — N{record['unit_price']} | Total: N{record['running_total']}")


def log_session_summary(summary: dict):
    """
    Logs the full session in two places:
      1. A new row in the Purchase History sheet (one row per session).
      2. A merged, bolded closing total row appended directly under
         that customer's item rows in the Item Events sheet — so
         scrolling through Item Events shows each session's items
         followed immediately by its total, not just a running
         number with no visible "session ended here" marker.
    summary: dict from LiveSession.get_session_summary()
    """
    if not os.path.exists(EXCEL_FILE):
        create_excel_file()

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # ── 1. Purchase History row (unchanged) ──
    ws = wb["Purchase History"]
    next_row = ws.max_row + 1
    session_id = f"SES{next_row - 1:04d}"

    # Only show items that were taken (not returned)
    taken_items = [b for b in summary["billed_items"] if b["direction"] == "taken"]
    items_str = ", ".join([
        f"{b['item']} (N{b['unit_price']})"
        for b in taken_items
    ])

    row_colour = "fff0f0" if next_row % 2 == 0 else "ffffff"
    row_fill = PatternFill("solid", fgColor=row_colour)

    row_data = [
        session_id,
        summary["start_time"],
        summary["end_time"],
        summary["customer_name"],
        summary["staff_id"],
        items_str,
        summary["total"]
    ]

    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="left")
        cell.font = Font(size=10)

    ws.cell(row=next_row, column=7).font = Font(bold=True, size=10)

    # ── 2. Closing total row in Item Events (new) ──
    ws_events = wb["Item Events"]
    event_total_row = ws_events.max_row + 1
    summary_text = (
        f"SESSION TOTAL — {summary['customer_name']} "
        f"({summary['staff_id']}) — {session_id}: N{summary['total']}"
    )
    ws_events.merge_cells(
        start_row=event_total_row, start_column=1,
        end_row=event_total_row, end_column=len(EVENT_HEADERS)
    )
    sc = ws_events.cell(event_total_row, 1, summary_text)
    sc.font = Font(bold=True, size=10, color="FFFFFF")
    sc.fill = PatternFill("solid", fgColor="2E5E3E")
    sc.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(EXCEL_FILE)
    print(f"Purchase history logged: {session_id} — {summary['customer_name']} — N{summary['total']}")